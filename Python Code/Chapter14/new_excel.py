from openpyxl import load_workbook

wb = load_workbook("Chapter14/Dodgers.xlsx")
result = []

ws = wb.worksheets[0]
for row in ws.iter_rows():
    result.append([cell.value for cell in row])
print(result)
